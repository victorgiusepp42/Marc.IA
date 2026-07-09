"""
Gera planilha XLSX com todas as tarefas do Marc.IA.
Saída: docs/MarcIA_Tarefas_2026-07-01.xlsx

Uso: python scripts/gerar_planilha_tarefas.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === Dados: (Sprint, ID, Tarefa, Linguagem, Stack, Complexidade, Deps, Responsavel, Status) ===
TAREFAS = [
    # ---- SPRINT 1 ----
    ("S1", "FUND-01", "venv + .gitignore (venv/, .env, __pycache__/, data/sessoes/)", "Config", "raiz", "S", "—", "V", "✅"),
    ("S1", "FUND-02", "requirements.txt (flask, anthropic, python-dotenv, flask-sqlalchemy, chromadb, sentence-transformers, pymupdf)", "Config", "raiz", "S", "—", "V", "✅"),
    ("S1", "DEC-01", "Definir modelo API (Opus 4.8 primário + Sonnet 4.6 fallback)", "Linguagem natural", "decisão", "S", "—", "V", "✅"),
    ("S1", "DEC-02", "ANTHROPIC_API_KEY no .env + ANTHROPIC_API_KEY_FALLBACK", "Config", ".env", "S", "FUND-01", "V", "✅"),
    ("S1", "NUC-01", "Classe Sessao (versão antiga JSON, substituída em NUC-07)", "Python", "src/sessao.py", "M", "—", "V", "✅"),
    ("S1", "NUC-02", "app.py mínimo (GET /, POST /chat)", "Python", "app.py", "S", "FUND-02", "V", "✅"),
    ("S1", "AGE-01", "AgentBase (cliente Anthropic, timeout 60s, retry)", "Python", "src/agente_base.py", "M", "FUND-02, DEC-02", "V", "✅"),
    ("S1", "AGE-02", "AgenteIP (persona PyAI Professor completa, system_prompt ~190 linhas)", "Python", "src/agente_ip.py", "L", "AGE-01", "V", "✅"),
    ("S1", "AGE-03", "Integração app↔agente (POST /chat chama AgenteIP.responder)", "Python", "app.py", "S", "NUC-02, AGE-02, NUC-01", "V", "✅"),
    ("S1", "CONT-01", "data/plano.json com 10 tópicos de IP", "JSON", "data/", "S", "—", "N", "✅"),
    ("S1", "FRO-01", "index.html (header, sidebar rota+XP, chat-area, form)", "HTML", "templates/", "S", "—", "D", "✅"),
    ("S1", "FRO-02", "style.css 8-bit (paleta azul/amarelo/fundo, Press Start 2P, sem anti-aliasing, responsivo básico)", "CSS", "static/", "M", "FRO-01", "D", "✅"),
    ("S1", "FRO-03", "chat.js (fetch, render, double-submit, pensando..., timeout 45s via AbortController)", "JS", "static/", "M", "FRO-01, NUC-02", "D", "✅"),

    # ---- SPRINT 2 - Bloco A: Banco ----
    ("S2-A", "NUC-06", "SQLAlchemy setup (database.py, DATABASE_URL, db.create_all, *.db no .gitignore)", "Python+ORM", "src/database.py", "M", "FUND-02", "K", "✅"),
    ("S2-A", "NUC-07", "Migrar Sessao para banco (db.Model, historico_json, factory carregar_ou_criar)", "Python+ORM", "src/models.py", "L", "NUC-01, NUC-06", "K", "✅"),
    ("S2-A", "NUC-05", "Progresso/XP/nível (lógica, integrar em Sessao.adicionar_xp, recalcular nível)", "Python", "src/models.py (XP_NIVEIS)", "M", "NUC-01, NUC-07", "V", "✅"),
    ("S2-A", "NUC-04", "Rota loader (GET /rota, extrair lógica para src/rota_loader.py)", "Python", "src/rota_loader.py, app.py", "S", "CONT-01", "V", "🟡"),

    # ---- SPRINT 2 - Bloco B: Auth ----
    ("S2-B", "DEC-03", "Google Cloud OAuth credentials (projeto, consent screen, Client ID/Secret, redirect URI)", "Web UI externa", "console.cloud.google.com", "S", "—", "K", "⏳"),
    ("S2-B", "AUTH-01", "OAuth config (src/auth.py com Authlib, scopes openid email profile)", "Python", "src/auth.py", "M", "DEC-03, NUC-02", "K", "⏳"),
    ("S2-B", "AUTH-02", "Flask-Login + Authlib (rotas /login, /login/callback, /logout, @login_required)", "Python", "src/auth.py, app.py", "L", "AUTH-01, NUC-07", "K", "⏳"),
    ("S2-B", "AUTH-03", "Sessão por google_id (extrair sub, criar/carregar Usuario, vincular SessaoEstudo)", "Python+ORM", "src/auth.py", "M", "AUTH-02, NUC-07", "K", "⏳"),

    # ---- SPRINT 2 - Bloco C: RAG ----
    ("S2-C", "CONT-03", "KB materiais IP (scripts/build_kb.py com pymupdf, ~100 TXTs extraídos)", "Python, Shell", "data/kb_ip/, scripts/build_kb.py", "L", "—", "N", "✅"),
    ("S2-C", "RAG-01", "src/rag/indexador.py (chunk ~2000 chars + overlap, embeddings, ChromaDB)", "Python", "src/rag/indexador.py", "L", "CONT-03", "N", "✅"),
    ("S2-C", "RAG-02", "src/rag/retriever.py (singleton, top-k=3, retorna [fonte]\\ntexto)", "Python", "src/rag/retriever.py", "M", "RAG-01", "N", "✅"),
    ("S2-C", "RAG-03", "Integração RAG↔Agente (AgentBase.buscar_contexto, chunks no prompt)", "Python", "src/agente_base.py", "S", "RAG-02, AGE-04", "V", "✅"),

    # ---- SPRINT 2 - Bloco D: Exercícios + Navegação ----
    ("S2-D", "CONT-02", "data/exercicios.json com 5+ exercícios por tópico (multipla_escolha, resposta_curta, codigo)", "JSON", "data/exercicios.json", "L", "CONT-01", "N", "⏳"),
    ("S2-D", "EXE-01", "src/exercicios.py (carregar_exercicios, selecionar_exercicio, validar_resposta)", "Python", "src/exercicios.py", "M", "CONT-02", "N", "⏳"),
    ("S2-D", "AGE-04", "Histórico no prompt (_montar_messages com últimas 10 mensagens)", "Python", "src/agente_base.py", "S", "AGE-01, NUC-03", "V", "✅"),
    ("S2-D", "AGE-05", "Prompt adaptativo dinâmico (interpolar topico_ativo no system_prompt)", "Python, Linguagem natural", "src/agente_ip.py", "M", "AGE-02, NUC-04", "V", "🟡"),
    ("S2-D", "FRO-04", "Painel rota dinâmico (fetch /rota, render <li> com ícones)", "JS", "static/chat.js", "S", "NUC-04, FRO-01", "D", "🟡"),
    ("S2-D", "FRO-04b", "Navegação livre (POST /rota/selecionar, atualiza topico_ativo)", "JS", "static/chat.js, app.py", "M", "FRO-04", "D", "⏳"),
    ("S2-D", "FRO-05", "Exercício inline (card, POST /exercicio/validar, feedback animado)", "JS, Python", "static/chat.js, app.py, src/exercicios.py", "L", "EXE-01", "D", "⏳"),

    # ---- SPRINT 3 - Bloco A: Métricas ----
    ("S3-A", "MET-01", "Feedback por resposta (botões like/dislike, POST /feedback, model Feedback)", "Python+ORM, JS", "static/chat.js, app.py, src/models.py", "M", "NUC-07, FRO-03", "K", "⏳"),
    ("S3-A", "MET-02", "NPS + Sean Ellis (modal a cada 5 sessões, POST /metricas, model Metrica)", "Python+ORM, JS", "static/chat.js, app.py, src/models.py", "M", "NUC-07", "K", "⏳"),
    ("S3-A", "FRO-06", "UI de feedback (botões pixel art, modal NPS, slider custom, X de fechar)", "CSS, HTML, JS", "templates/index.html, static/style.css, static/chat.js", "M", "MET-01, MET-02", "D", "⏳"),

    # ---- SPRINT 3 - Bloco B: QA ----
    ("S3-B", "QA-01", "pytest + conftest (fixtures app, client, mock_claude, pytest.ini)", "Python", "tests/, pytest.ini", "M", "AUTH-02", "N", "⏳"),
    ("S3-B", "QA-02", "Testes unitários (test_sessao, test_progresso, test_exercicios, test_validacao, test_rotas, meta ≥50%)", "Python", "tests/", "L", "QA-01, EXE-01", "N", "⏳"),
    ("S3-B", "QA-03", "src/validacao.py (sanitizar, validar_mensagem, integrar em POST /chat)", "Python", "src/validacao.py, app.py", "M", "NUC-02", "N", "🟡"),

    # ---- SPRINT 3 - Bloco C: Deploy ----
    ("S3-C", "MOS-01", "Deploy Railway (criar projeto, conectar repo, Procfile, ler PORT do env)", "Config, Python", "Procfile, app.py", "M", "K", "K", "⏳"),
    ("S3-C", "MOS-02", "PostgreSQL Railway (addon, DATABASE_URL auto-injetado, validar db.create_all)", "Python+ORM", "railway dashboard", "S", "MOS-01, NUC-06", "K", "⏳"),
    ("S3-C", "MOS-03", "Variáveis ambiente Railway (ANTHROPIC_API_KEY, GOOGLE_*, SECRET_KEY, DATABASE_URL) + redirect URI prod", "Config, Web UI externa", "railway dashboard + google cloud", "S", "MOS-01, DEC-03", "K", "⏳"),

    # ---- SPRINT 3 - Bloco D: Docs ----
    ("S3-D", "DOC-01", "Relatório E2 (PDF 1 página: problema, solução, arquitetura, status, screenshot)", "Linguagem natural", "docs/relatorio_e2.pdf", "M", "—", "V", "🟡"),
    ("S3-D", "DOC-02", "Vídeo E2 (1-3min mostrando app funcional)", "Linguagem natural", "—", "M", "—", "T", "⏳"),
    ("S3-D", "DOC-03", "README final (Sobre, Instalação, Uso, Arquitetura, Tech Stack, Equipe, screenshots)", "Linguagem natural", "README.md", "M", "—", "V", "🟡"),
    ("S3-D", "DOC-04", "Documentação técnica (3-5 páginas, diagramas Mermaid renderizados, PDF acadêmico)", "Linguagem natural", "docs/doc_tecnica.pdf", "L", "SDD.md", "V", "🟡"),
    ("S3-D", "DOC-05", "Pôster A1 (594×841mm, paleta 8-bit, seções completas, QR code)", "Linguagem natural", "docs/poster_a1.pdf", "L", "—", "T", "⏳"),
    ("S3-D", "DOC-06", "Vídeo final (2-5min, demo completa via Railway)", "Linguagem natural", "—", "M", "MOS-01", "T", "⏳"),

    # ---- SPRINT 3 - Bloco E: Polimento ----
    ("S3-E", "FRO-07", "Responsivo mobile (media queries 360px/768px, hambúrguer)", "CSS", "static/style.css", "M", "FRO-02", "D", "🟡"),
    ("S3-E", "FRO-08", "Loading states (bolha '...' piscando, desabilita botão enviar)", "JS", "static/chat.js, static/style.css", "S", "FRO-03", "D", "✅"),
    ("S3-E", "POL-01", "Tratamento de erros completo (@app.errorhandler 500/404, logging.error, sem 500 sem msg)", "Python", "src/agente_base.py, app.py, src/auth.py", "M", "AGE-01", "V", "🟡"),
    ("S3-E", "POL-02", "Garantir while em código real (já existe em indexador.py linha 31)", "Python", "src/rag/indexador.py", "XS", "—", "V", "✅"),
    ("S3-E", "POL-03", "Garantir tupla em código real (já existe XP_NIVEIS em models.py)", "Python", "src/models.py", "XS", "—", "V", "✅"),

    # ---- EXTRAS (raio-x) ----
    ("EXTRA", "EXTRA-01", "Commit de tudo untracked (app.py, src/, data/, docs/, templates/, static/, scripts/, requirements.txt, .gitignore, CONTEXTO_PROJETO.md)", "Shell", "terminal", "S", "—", "V", "⏳"),
    ("EXTRA", "EXTRA-02", "Backup Drive (copiar para Marc.IA_Projeto_IP/backup_github/)", "Shell", "terminal", "S", "—", "V", "⏳"),
    ("EXTRA", "EXTRA-03", "Limpar data/sessoes/*.json e instance/marcia.db (sessões antigas)", "Shell", "filesystem", "XS", "—", "V", "⏳"),
    ("EXTRA", "EXTRA-04", "Verificar se src/sessao.py ainda é usado (foi substituído por models.py)", "Python", "src/sessao.py", "XS", "—", "V", "⏳"),
    ("EXTRA", "EXTRA-05", "Criar Procfile (mencionado em MOS-01 mas não existe)", "Config", "raiz", "XS", "—", "K", "⏳"),
    ("EXTRA", "EXTRA-06", "Criar pasta tests/ com .gitkeep", "Config", "tests/", "XS", "—", "N", "⏳"),
    ("EXTRA", "EXTRA-07", "Garantir instance/ no .gitignore (SQLite)", "Config", ".gitignore", "XS", "—", "V", "⏳"),
]


def main():
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    docs_dir = os.path.join(base, "docs")
    os.makedirs(docs_dir, exist_ok=True)

    out_path = os.path.join(docs_dir, "MarcIA_Tarefas_2026-07-01.xlsx")

    wb = Workbook()

    # ==================== ABA 1: Visão geral ====================
    ws = wb.active
    ws.title = "Tarefas"

    headers = ["Sprint", "ID", "Tarefa", "Linguagem", "Stack", "Complexidade", "Deps", "Resp", "Status"]

    # Estilo header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A3F6B", end_color="1A3F6B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="2B6CB0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Status → cor de fundo
    status_fill = {
        "✅": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "🟡": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "⏳": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "❌": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    # Complexidade → cor
    comp_fill = {
        "XS": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "S":  PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "M":  PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "L":  PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "XL": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    }

    body_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, row in enumerate(TAREFAS, start=2):
        sprint, tid, tarefa, lang, stack, comp, deps, resp, status = row
        cells = [sprint, tid, tarefa, lang, stack, comp, deps, resp, status]
        for col_idx, val in enumerate(cells, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = body_align
            cell.border = border
        # pinta status
        c = ws.cell(row=row_idx, column=9)
        c.fill = status_fill.get(status, PatternFill())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=12)
        # pinta complexidade
        c = ws.cell(row=row_idx, column=6)
        c.fill = comp_fill.get(comp, PatternFill())
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True)
        # pinta responsável
        c = ws.cell(row=row_idx, column=8)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True)
        # pinta sprint
        c = ws.cell(row=row_idx, column=1)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, color="2B6CB0")

    # Larguras
    widths = [9, 12, 70, 30, 40, 13, 25, 8, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Linha header mais alta
    ws.row_dimensions[1].height = 30

    # Congela painel
    ws.freeze_panes = "A2"

    # Auto-filtro
    ws.auto_filter.ref = ws.dimensions

    # ==================== ABA 2: Resumo ====================
    ws2 = wb.create_sheet("Resumo")

    total = len(TAREFAS)
    feito = sum(1 for r in TAREFAS if r[8] == "✅")
    parcial = sum(1 for r in TAREFAS if r[8] == "🟡")
    pendente = sum(1 for r in TAREFAS if r[8] == "⏳")

    # Por linguagem
    from collections import Counter
    langs = []
    for r in TAREFAS:
        for lang in r[3].split(","):
            langs.append(lang.strip())
    lang_counter = Counter(langs)

    # Por responsável
    resp_counter = Counter(r[7] for r in TAREFAS)

    # Por sprint
    sprint_counter = Counter(r[0] for r in TAREFAS)

    # Por complexidade
    comp_counter = Counter(r[5] for r in TAREFAS)

    ws2.cell(row=1, column=1, value="Marc.IA — Resumo de Tarefas").font = Font(bold=True, size=14, color="1A3F6B")
    ws2.cell(row=2, column=1, value=f"Gerado em 01/07/2026 — Total: {total} tarefas").font = Font(italic=True)

    # Status geral
    ws2.cell(row=4, column=1, value="STATUS GERAL").font = Font(bold=True, size=12, color="FFFFFF")
    ws2.cell(row=4, column=1).fill = header_fill
    ws2.cell(row=4, column=2).fill = header_fill
    ws2.cell(row=4, column=1).alignment = header_align
    ws2.cell(row=4, column=2).alignment = header_align

    status_rows = [
        ("✅ Concluído", feito),
        ("🟡 Parcial", parcial),
        ("⏳ Pendente", pendente),
        ("Total", total),
        ("% Concluído", f"{feito*100//total}%"),
    ]
    for i, (label, val) in enumerate(status_rows, start=5):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws2.cell(row=i, column=2, value=val)
        c.alignment = Alignment(horizontal="center")
        if label == "✅ Concluído":
            c.fill = status_fill["✅"]
        elif label == "🟡 Parcial":
            c.fill = status_fill["🟡"]
        elif label == "⏳ Pendente":
            c.fill = status_fill["⏳"]

    # Por linguagem
    row_atual = 12
    ws2.cell(row=row_atual, column=1, value="POR LINGUAGEM").font = Font(bold=True, size=12, color="FFFFFF")
    ws2.cell(row=row_atual, column=1).fill = header_fill
    ws2.cell(row=row_atual, column=2).fill = header_fill
    ws2.cell(row=row_atual, column=3).fill = header_fill
    for col in range(1, 4):
        ws2.cell(row=row_atual, column=col).alignment = header_align
    ws2.cell(row=row_atual, column=1, value="Linguagem")
    ws2.cell(row=row_atual, column=2, value="Qtd")
    ws2.cell(row=row_atual, column=3, value="%")
    row_atual += 1
    for lang, qtd in sorted(lang_counter.items(), key=lambda x: -x[1]):
        ws2.cell(row=row_atual, column=1, value=lang).font = Font(bold=True)
        c = ws2.cell(row=row_atual, column=2, value=qtd)
        c.alignment = Alignment(horizontal="center")
        ws2.cell(row=row_atual, column=3, value=f"{qtd*100//total}%").alignment = Alignment(horizontal="center")
        row_atual += 1

    # Por responsável
    row_atual += 1
    ws2.cell(row=row_atual, column=1, value="POR RESPONSÁVEL").font = Font(bold=True, size=12, color="FFFFFF")
    ws2.cell(row=row_atual, column=1).fill = header_fill
    ws2.cell(row=row_atual, column=2).fill = header_fill
    ws2.cell(row=row_atual, column=3).fill = header_fill
    for col in range(1, 4):
        ws2.cell(row=row_atual, column=col).alignment = header_align
    ws2.cell(row=row_atual, column=1, value="Responsável")
    ws2.cell(row=row_atual, column=2, value="Qtd")
    ws2.cell(row=row_atual, column=3, value="%")
    row_atual += 1
    resp_nomes = {"V": "Victor", "D": "Diogo", "K": "Kevynson", "N": "Natan", "T": "Todos"}
    for resp, qtd in sorted(resp_counter.items(), key=lambda x: -x[1]):
        ws2.cell(row=row_atual, column=1, value=f"{resp} — {resp_nomes.get(resp, '?')}").font = Font(bold=True)
        c = ws2.cell(row=row_atual, column=2, value=qtd)
        c.alignment = Alignment(horizontal="center")
        ws2.cell(row=row_atual, column=3, value=f"{qtd*100//total}%").alignment = Alignment(horizontal="center")
        row_atual += 1

    # Por sprint
    row_atual += 1
    ws2.cell(row=row_atual, column=1, value="POR SPRINT").font = Font(bold=True, size=12, color="FFFFFF")
    ws2.cell(row=row_atual, column=1).fill = header_fill
    ws2.cell(row=row_atual, column=2).fill = header_fill
    ws2.cell(row=row_atual, column=3).fill = header_fill
    for col in range(1, 4):
        ws2.cell(row=row_atual, column=col).alignment = header_align
    ws2.cell(row=row_atual, column=1, value="Sprint")
    ws2.cell(row=row_atual, column=2, value="Qtd")
    ws2.cell(row=row_atual, column=3, value="%")
    row_atual += 1
    for sprint, qtd in sorted(sprint_counter.items()):
        ws2.cell(row=row_atual, column=1, value=sprint).font = Font(bold=True)
        c = ws2.cell(row=row_atual, column=2, value=qtd)
        c.alignment = Alignment(horizontal="center")
        ws2.cell(row=row_atual, column=3, value=f"{qtd*100//total}%").alignment = Alignment(horizontal="center")
        row_atual += 1

    # Por complexidade
    row_atual += 1
    ws2.cell(row=row_atual, column=1, value="POR COMPLEXIDADE").font = Font(bold=True, size=12, color="FFFFFF")
    ws2.cell(row=row_atual, column=1).fill = header_fill
    ws2.cell(row=row_atual, column=2).fill = header_fill
    ws2.cell(row=row_atual, column=3).fill = header_fill
    for col in range(1, 4):
        ws2.cell(row=row_atual, column=col).alignment = header_align
    ws2.cell(row=row_atual, column=1, value="Complexidade")
    ws2.cell(row=row_atual, column=2, value="Qtd")
    ws2.cell(row=row_atual, column=3, value="%")
    row_atual += 1
    ordem_comp = ["XS", "S", "M", "L", "XL"]
    for comp in ordem_comp:
        qtd = comp_counter.get(comp, 0)
        ws2.cell(row=row_atual, column=1, value=comp).font = Font(bold=True)
        c = ws2.cell(row=row_atual, column=2, value=qtd)
        c.alignment = Alignment(horizontal="center")
        c.fill = comp_fill.get(comp, PatternFill())
        ws2.cell(row=row_atual, column=3, value=f"{qtd*100//total}%" if total else "0%").alignment = Alignment(horizontal="center")
        row_atual += 1

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    # ==================== ABA 3: Legenda ====================
    ws3 = wb.create_sheet("Legenda")
    legenda = [
        ["Campo", "Significado"],
        ["Sprint", "S1 = Walking Skeleton · S2-A = Banco · S2-B = Auth · S2-C = RAG · S2-D = Exercícios · S3-A = Métricas · S3-B = QA · S3-C = Deploy · S3-D = Docs · S3-E = Polimento · EXTRA = Tarefas extras"],
        ["ID", "Identificador da tarefa (Épico-Número)"],
        ["Tarefa", "Descrição resumida do que deve ser feito"],
        ["Linguagem", "Stack técnico da tarefa: Python, HTML, CSS, JS, JSON, Config, Linguagem natural, Python+ORM, Shell, Web UI externa"],
        ["Stack", "Arquivo/pasta onde mexer"],
        ["Complexidade", "XS = ≤30min · S = ≤2h · M = 2-6h · L = 6-16h · XL = >16h"],
        ["Deps", "Tarefas que precisam estar concluídas antes"],
        ["Resp", "V = Victor · D = Diogo · K = Kevynson · N = Natan · T = Todos"],
        ["Status", "✅ = Concluído · 🟡 = Parcial · ⏳ = Pendente · ❌ = Bloqueado"],
    ]
    for r_idx, row in enumerate(legenda, start=1):
        for c_idx, val in enumerate(row, start=1):
            c = ws3.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            if r_idx == 1:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = header_fill
                c.alignment = header_align
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 110
    ws3.row_dimensions[1].height = 24

    # Salva
    wb.save(out_path)
    print(f"[OK] Planilha gerada: {out_path}")
    print(f"     Total de tarefas: {total}")
    print(f"     Concluidas:  {feito} ({feito*100//total}%)")
    print(f"     Parciais:    {parcial}")
    print(f"     Pendentes:   {pendente}")


if __name__ == "__main__":
    main()